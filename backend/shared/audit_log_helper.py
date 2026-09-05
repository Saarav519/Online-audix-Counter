"""
Audit Log helper — shared between warehouse (audit_routes.py) and cycle
count (cycle_count_routes.py). Implements the "Movement / Audit Log"
feature: a single `audit_logs` MongoDB collection that captures every
mutation (edit, undo, reco_adjust, delete, assign, revoke) across both
modules with consistent fields.

Design rules (per spec):

  • Logging is async + non-blocking. If `log_audit_entry` fails for any
    reason, the caller's edit MUST still succeed. Wrapped in
    try/except → never raises.

  • One shared schema across modules. Module column distinguishes
    warehouse vs cycle_count entries.

  • All filter / search / export logic lives here so audit_routes.py
    and cycle_count_routes.py just call into the helper.
"""

from __future__ import annotations

import io
import logging
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# Allowed action types (kept loose so callers can extend without code
# changes — but documented here for downstream filters / UI badges).
ACTION_TYPES = {"edit", "undo", "reco_adjust", "delete", "assign", "revoke"}
MODULES = {"warehouse", "cycle_count"}


def _iso(dt: Optional[datetime] = None) -> str:
    return (dt or datetime.now(timezone.utc)).isoformat()


def _norm(v: Any) -> str:
    """Stringify a value for storage — keep None as empty string so the
    UI doesn't render 'None'. Numbers stay as their string repr."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        # Drop .0 for whole-number floats so the UI reads cleanly.
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    return str(v)


async def log_audit_entry(db, entry_data: Dict[str, Any]) -> None:
    """Insert a single audit-trail entry. Fire-and-forget semantics: a
    failure here NEVER bubbles to the caller. The caller's edit must
    succeed even if logging fails.

    Required-ish fields (any missing defaulted to empty / sane value):
      module, action_type, barcode, client_id, old_value, new_value
    Optional:
      session_id, cycle_day, field_name, user_id, username, report_type,
      location
    """
    try:
        now = datetime.now(timezone.utc)
        entry = {
            "id": str(uuid.uuid4()),
            "module": entry_data.get("module") or "warehouse",
            "action_type": entry_data.get("action_type") or "edit",
            "barcode": _norm(entry_data.get("barcode")),
            "client_id": entry_data.get("client_id") or "",
            "session_id": entry_data.get("session_id") or "",
            "cycle_day": entry_data.get("cycle_day"),  # int | None
            "field_name": entry_data.get("field_name") or "barcode",
            "old_value": _norm(entry_data.get("old_value")),
            "new_value": _norm(entry_data.get("new_value")),
            "user_id": entry_data.get("user_id") or "",
            "username": entry_data.get("username") or "",
            "report_type": entry_data.get("report_type") or "",
            "location": entry_data.get("location") or "",
            "timestamp": _iso(now),
            # Keep a true datetime as well so we can $gte/$lte by date
            # directly without ISO-string lexicographic gotchas.
            "timestamp_dt": now,
        }
        await db.audit_logs.insert_one(entry)
    except Exception as e:
        # NEVER raise — audit logging must not break the caller's edit.
        logger.warning(f"audit_logs insert failed (non-fatal): {e}")


def _build_search_query(filters: Dict[str, Any]) -> Dict[str, Any]:
    """Translate the user-facing filter payload into a MongoDB query.
    Empty / falsy filters are skipped so the user can leave fields blank.
    """
    q: Dict[str, Any] = {}
    if filters.get("module") and filters["module"] != "all":
        q["module"] = filters["module"]
    cid = filters.get("client_id")
    if isinstance(cid, (list, tuple, set)):
        # A list arrives when the caller is scoped to a set of clients. An empty
        # one means "nothing visible" and must match no rows — not everything.
        q["client_id"] = {"$in": list(cid)}
    elif cid:
        q["client_id"] = cid
    if filters.get("session_id"):
        q["session_id"] = filters["session_id"]
    if filters.get("user_id"):
        q["user_id"] = filters["user_id"]
    if filters.get("cycle_day") not in (None, ""):
        try:
            q["cycle_day"] = int(filters["cycle_day"])
        except (TypeError, ValueError):
            pass
    if filters.get("barcode"):
        # Substring match (case-insensitive) — auditors usually paste
        # partial codes.
        q["barcode"] = {"$regex": str(filters["barcode"]), "$options": "i"}
    if filters.get("action_type"):
        q["action_type"] = filters["action_type"]
    # Date range — accept ISO strings OR datetime objects from caller.
    start = filters.get("start_date")
    end = filters.get("end_date")
    if start or end:
        rng: Dict[str, Any] = {}
        if start:
            rng["$gte"] = _parse_date_floor(start)
        if end:
            rng["$lte"] = _parse_date_ceil(end)
        if rng:
            q["timestamp_dt"] = rng
    return q


def _parse_date_floor(v: Any) -> datetime:
    """Parse user-supplied date → start-of-day UTC."""
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    s = str(v)
    # Accept full ISO; fall back to YYYY-MM-DD = start of day
    try:
        d = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
    except Exception:
        try:
            d = datetime.strptime(s[:10], "%Y-%m-%d")
            return d.replace(tzinfo=timezone.utc)
        except Exception:
            return datetime(1970, 1, 1, tzinfo=timezone.utc)


def _parse_date_ceil(v: Any) -> datetime:
    """Parse user-supplied date → end-of-day UTC (so inclusive end works).

    NB: `datetime.fromisoformat("2026-05-26")` succeeds in Python 3.11+
    and yields midnight. For bare-date inputs we still want
    end-of-day, so we detect "no time component supplied" via the
    string length (10 chars = YYYY-MM-DD) and bump to 23:59:59.999.
    """
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    s = str(v).strip()
    bare_date = len(s) == 10  # 'YYYY-MM-DD'
    try:
        d = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if bare_date:
            d = d.replace(hour=23, minute=59, second=59, microsecond=999000)
        return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
    except Exception:
        try:
            d = datetime.strptime(s[:10], "%Y-%m-%d")
            return d.replace(hour=23, minute=59, second=59, microsecond=999000, tzinfo=timezone.utc)
        except Exception:
            return datetime(9999, 12, 31, tzinfo=timezone.utc)


async def search_audit_logs(db, filters: Dict[str, Any]) -> Dict[str, Any]:
    """Paginated search over `audit_logs`. Returns {logs, total}.

    Pagination:
      limit (default 50, capped 500), skip (default 0).
    Sort:
      timestamp desc (most recent first).
    """
    limit = max(1, min(int(filters.get("limit") or 50), 500))
    skip = max(0, int(filters.get("skip") or 0))
    q = _build_search_query(filters)

    cursor = db.audit_logs.find(q, {"_id": 0, "timestamp_dt": 0}).sort("timestamp_dt", -1).skip(skip).limit(limit)
    logs = await cursor.to_list(limit)
    total = await db.audit_logs.count_documents(q)
    return {"logs": logs, "total": total, "limit": limit, "skip": skip}


async def fetch_recent_for_barcode(db, *, barcode: str, client_id: Optional[str] = None,
                                   limit: int = 5) -> List[Dict[str, Any]]:
    """Last N audit-log entries for a given barcode (and optionally a
    client). Used by the "Last Edited" popup in both modules.
    """
    if not barcode:
        return []
    q: Dict[str, Any] = {"barcode": barcode}
    if client_id:
        q["client_id"] = client_id
    cursor = db.audit_logs.find(q, {"_id": 0, "timestamp_dt": 0}).sort("timestamp_dt", -1).limit(max(1, min(limit, 50)))
    return await cursor.to_list(50)


async def export_audit_logs_excel(db, filters: Dict[str, Any]) -> bytes:
    """Export ALL matching audit logs (ignoring pagination) as an Excel
    file. Returns raw bytes — the FastAPI route wraps them in a
    StreamingResponse.

    Hard cap of 50,000 rows so a runaway export doesn't OOM the worker.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    q = _build_search_query(filters)
    cursor = db.audit_logs.find(q, {"_id": 0, "timestamp_dt": 0}).sort("timestamp_dt", -1).limit(50000)
    rows = await cursor.to_list(50000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"

    headers = [
        "Timestamp", "Module", "Action", "User", "Client",
        "Session", "Day", "Barcode", "Field", "Old Value",
        "New Value", "Report Type", "Location",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([
            r.get("timestamp", ""),
            r.get("module", ""),
            r.get("action_type", ""),
            r.get("username", "") or r.get("user_id", ""),
            r.get("client_id", ""),
            r.get("session_id", ""),
            r.get("cycle_day", "") if r.get("cycle_day") is not None else "",
            r.get("barcode", ""),
            r.get("field_name", ""),
            r.get("old_value", ""),
            r.get("new_value", ""),
            r.get("report_type", ""),
            r.get("location", ""),
        ])

    # Auto-ish column widths
    widths = [22, 12, 12, 16, 20, 22, 6, 22, 14, 22, 22, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def resolve_module_for_client(db, client_id: Optional[str]) -> str:
    """Helper used by callers that don't already know whether they're
    operating in warehouse or cycle_count context — look up the client
    document and return the module string. Defaults to 'warehouse' if
    the client doesn't exist or has no client_type."""
    if not client_id:
        return "warehouse"
    try:
        c = await db.clients.find_one({"id": client_id}, {"_id": 0, "client_type": 1})
        if c and c.get("client_type") == "cycle_count":
            return "cycle_count"
    except Exception:
        pass
    return "warehouse"
